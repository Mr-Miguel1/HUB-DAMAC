import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

import os
import xlrd
import shutil


def clean_mlaboral_BR(path):
    try:
        os.mkdir(path+"\\CSV")
    except:
        pass 
    try:
        archivos = os.listdir(path)
        archivos.remove("CSV")
        for i in archivos:
            data = pd.read_csv(path+"\{}".format(i),sep=';',decimal=',')
            data.set_index('Fecha',drop=True,inplace=True)
            data = data.applymap(lambda x: float(x)/100)
            os.remove(path+"\{}".format(i))
            data.to_csv(path+"\CSV\{}.csv".format(i[:-4]),sep=';',decimal=',')
    except:
        print("Los datos de BanRep no se limpiaron correctamente")
        pass

def clean_informalidad(path):
    try:
        os.mkdir(path+"\\archivos_fuente")
        os.mkdir(path+"\\CSV")
    except:
        pass 
    try:
        archivos =  os.listdir(path)
        archivos.remove("archivos_fuente")
        archivos.remove("CSV")
        dane_informalidad_nombre = archivos[0]
        for i in archivos:
            if i ==  dane_informalidad_nombre:
                try:
                    data = load_workbook(path+"\{}".format(i))
                    sheets = data.sheetnames
                except:
                    data = xlrd.open_workbook_xls(path+"\{}".format(i))
                    sheets = sheets = data.sheet_names()


                #Prop informalidad Total Nacional
                try:
                    sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                    tnal_nacional_informalidad_index = sheets[sheets.str.contains('informalidad')].index[0]

                    df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_informalidad_index)

                    index = df[df.iloc[:,0].str.contains('23 Ciudades').fillna(False)].index[0]

                    data_23_ciudades = df.iloc[index:,:]

                    periodo = data_23_ciudades[data_23_ciudades.iloc[:,1].str.contains('Ene-').fillna(False)].dropna(axis=1).T
                    periodo.columns = ['Periodo']
                    periodo.reset_index(level=0,drop=True,inplace = True)

                    fecha = pd.date_range(start='2007',freq='M',periods=len(periodo),name = 'Fecha')
                    periodo['Fecha'] = fecha

                    tasa = data_23_ciudades[data_23_ciudades.iloc[:,0].str.contains('Ocupados').fillna(False)].dropna(axis=1).T.iloc[1:,:]

                    tasa.columns = ['Ocupados Informales']
                    tasa['Ocupados Informales'] = tasa['Ocupados Informales'].astype('float')/100
                    tasa.reset_index(level=0,drop=True,inplace=True)

                    tasa_informalidad = pd.concat([periodo,tasa],axis=1)
                    tasa_informalidad.to_csv(path+'\CSV\{}_total_Nacional.csv'.format(dane_informalidad_nombre[:-5]),sep=';',decimal=',',index=False)
                except:
                    print('La propoción de informalidad total nacional no se pudo limpiar correctamente')
                    pass


                #Informalidad Ciudad
                try:
                    sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                    tnal_nacional_ciudad_index = sheets[sheets.str.contains('ciudades')].index[0]

                    df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_ciudad_index)

                    l = df.iloc[:,0]
                    sup = l[l.str.contains('Total 13 ciudades y AM').fillna(False)].index[-1]
                    inf = l[l.str.contains('23 ciudades y áreas').fillna(False)].index[-1]

                    df = df.iloc[sup:inf+5,:]

                    df.reset_index(level=0,drop=True,inplace=True)

                    ind = [i.lower().replace(' ','_') for i in pd.Series("""Ocupados
                    Formales
                    Informales""").str.split('\n')[0]]

                    dic = pd.DataFrame({})
                    for j in ind:
                        ser_index_nac = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == 'informales')].dropna(how='all',axis=0).index
                    for ix in ser_index_nac:
                        ser_ = df.iloc[ix-4:ix+1,0:]
                        ser_ = ser_.T
                        ser_.columns = ser_.iloc[0,:]
                        ser_ = ser_.reset_index(level=0,drop=True).drop([0],axis=0)
                        fecha = pd.date_range(start='2007-01-01',periods=len(ser_),freq='M',name='Fecha')
                        ser_ = ser_.set_index(fecha,drop=True)
                        ser_['Ocupados'] = ser_['Ocupados'].astype('float')*1000
                        ser_['Formales'] = ser_['Formales'].astype('float')*1000
                        ser_['Informales'] = ser_['Informales'].astype('float')*1000

                        ser_ = ser_.rename(columns={'Ocupados':'Ocupados_{}'.format(ser_.columns[0][:3]),
                                                    'Formales':'Formales_{}'.format(ser_.columns[0][:3]),
                                                    'Informales':'Informales_{}'.format(ser_.columns[0][:3])
                                                    })

                        ser_ = ser_.fillna(method='ffill')

                        dic['Trimestre Móvil'] = ser_.iloc[:,1]

                        dic[ser_.columns[0]] = ser_.iloc[:,0].apply(lambda x: str(x).replace(str(x),ser_.columns[0]))
                        dic[ser_.columns[2]] = ser_.iloc[:,2].apply(lambda x: float(x))
                        dic[ser_.columns[3]] = ser_.iloc[:,3].apply(lambda x: float(x))
                        dic[ser_.columns[4]] = ser_.iloc[:,4].apply(lambda x: float(x))

                    dic.to_csv(path+r"\CSV\{}_ciudades.csv".format(dane_informalidad_nombre[:-5]),sep=';',decimal=',',encoding='utf-8')
                except:
                    print("La informalidad por Ciudades no se pudo limpiar correctamente")
                    pass

                #Informalidad por sexo
                try:
                    sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                    tnal_nacional_sexo_index = sheets[sheets.str.contains('sexo')].index[0]
                    df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_sexo_index)

                    ciudades = ["Ocupados 13 ciudades y áreas metropolitanas","Ocupados 23 ciudades y áreas metropolitanas"]

                    for j in ciudades:
                        l = df.iloc[:,0]
                        sup = l[l.str.contains(j).fillna(False)].index[0]

                        df_temp = df.iloc[sup:sup+9,:]
                        df_temp = df_temp.T.dropna(how='all',axis=0)
                        df_temp.columns = df_temp.iloc[0,:]

                        df_temp.reset_index(level=0,drop=True,inplace=True)
                        df_temp = df_temp.drop([0],axis=0)

                        fecha = pd.date_range(start='2007-01-01',periods=len(df_temp),freq='M',name='Fecha')

                        df_temp.set_index(fecha,drop=True,inplace=True)
                        df_temp = df_temp.applymap(lambda x: float(x)*1000)

                        df_temp.to_csv(path+r"\CSV\{}_sexo_{}.csv".format(dane_informalidad_nombre[:-5],j),sep=';',decimal=',',encoding='utf-8')
                except:
                    print("La informalidad por sexo no se pudo limpiar correctamente")
                    pass
                ## Imformalidad por educacion

                try:
                    sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                    tnal_nacional_educacion_index = sheets[sheets.str.contains('educación')].index[0]
                    df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_educacion_index)

                    ciudades = ["Ocupados 13 ciudades y áreas metropolitanas","Ocupados 23 ciudades y áreas metropolitanas"]

                    for j in ciudades:
                        l = df.iloc[:,0]
                        sup = l[l.str.contains(j).fillna(False)].index[0] 
                        df_temp = df.iloc[sup:sup+18,:]
                        df_temp = df_temp.T.dropna(how='all',axis=0)
                        df_temp.columns = df_temp.iloc[0,:]

                        df_temp.reset_index(level=0,drop=True,inplace=True)
                        df_temp = df_temp.drop([0],axis=0)

                        fecha = pd.date_range(start='2007-01-01',periods=len(df_temp),freq='M',name='Fecha')

                        df_temp.set_index(fecha,drop=True,inplace=True)
                        df_temp = df_temp.applymap(lambda x: float(x)*1000)

                        df_temp.to_csv(path+r"\CSV\{}_educacion_{}.csv".format(dane_informalidad_nombre[:-5],j),sep=';',decimal=',',encoding='utf-8')

                except:
                    print("La informalidad por educacion no se pudo limpiar correctamente")
                    pass


                ## Informalidad por ramas ciiu 4a
                try:
                    sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                    tnal_nacional_educacion_index = sheets[sheets.str.contains('ciiu4')].index[0]
                    df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_educacion_index)

                    ciudades = ["Total 13 áreas","Ocupados 23 ciudades y áreas metropolitanas"]

                    for j in ciudades:
                        l = df.iloc[:,0]
                        sup = l[l.str.contains(j).fillna(False)].index[0] 
                        df_temp = df.iloc[sup:sup+48,:]
                        df_temp = df_temp.T.dropna(how='all',axis=0)
                        df_temp.columns = df_temp.iloc[0,:]

                        df_temp.reset_index(level=0,drop=True,inplace=True)
                        df_temp = df_temp.drop([0],axis=0)

                        fecha = pd.date_range(start='2007-01-01',periods=len(df_temp),freq='M',name='Fecha')

                        df_temp.set_index(fecha,drop=True,inplace=True)
                        df_temp = df_temp.applymap(lambda x: float(x)*1000)

                        df_temp.to_csv(path+r"\CSV\{}_ramasciiu4a_{}.csv".format(dane_informalidad_nombre[:-5],j),sep=';',decimal=',',encoding='utf-8')
                except:
                    print("La informalidad por ramas CIIU4a no se pudo limpiar corectamente")
                    pass

                #Informalidad por Seguridad Social
                try:
                    sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                    tnal_nacional_educacion_index = sheets[sheets.str.contains('seguridadsocial13')].index[0]
                    df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_educacion_index)

                    ciudades = ["Ocupados 13 ciudades y áreas metropolitanas","Ocupados 23 ciudades y áreas metropolitanas"]


                    #Cantidad de personas
                    for j in ciudades:
                        l = df.iloc[:,0]
                        sup = l[l.str.contains(j).fillna(False)].index[0] 
                        df_temp = df.iloc[sup:sup+10,:]
                        df_temp = df_temp.T.dropna(how='all',axis=0)
                        df_temp.columns = df_temp.iloc[0,:]

                        df_temp.reset_index(level=0,drop=True,inplace=True)
                        df_temp = df_temp.drop([0],axis=0)

                        fecha = pd.date_range(start='2007-01-01',periods=len(df_temp),freq='M',name='Fecha')

                        df_temp.set_index(fecha,drop=True,inplace=True)
                        df_temp = df_temp.applymap(lambda x: float(x)*1000)

                        df_temp.to_csv(path+r"\CSV\{}_segsocial_cantidad_{}.csv".format(dane_informalidad_nombre[:-5],j),sep=';',decimal=',',encoding='utf-8')

                    #Porcentaje de personas
                    for j in ciudades:
                        l = df.iloc[:,0]
                        sup = l[l.str.contains(j).fillna(False)].index[1] 
                        df_temp = df.iloc[sup:sup+10,:]
                        df_temp = df_temp.T.dropna(how='all',axis=0)
                        df_temp.columns = df_temp.iloc[0,:]

                        df_temp.reset_index(level=0,drop=True,inplace=True)
                        df_temp = df_temp.drop([0],axis=0)

                        fecha = pd.date_range(start='2007-01-01',periods=len(df_temp),freq='M',name='Fecha')

                        df_temp.set_index(fecha,drop=True,inplace=True)
                        df_temp = df_temp.applymap(lambda x: float(x)/100)

                        df_temp.to_csv(path+r"\CSV\{}_segsocial_porcentaje_{}.csv".format(dane_informalidad_nombre[:-5],j),sep=';',decimal=',',encoding='utf-8')
                except:
                    print("La informalidad por seguridad social no se pudo limpiar correctamente")
                    pass

        shutil.move(path+r'\{}'.format(dane_informalidad_nombre),path+r'\archivos_fuente\{}'.format(dane_informalidad_nombre))
    except:
        pass

def clean_desempleo_desestacionalizado(path):
    
    try:
        os.mkdir(path+"\\archivos_fuente")
        os.mkdir(path+"\\CSV")
    except:
        pass 
    try:
        archivos =  os.listdir(path)
        archivos.remove("archivos_fuente")
        archivos.remove("CSV")
        dane_des_emp_mensual_nombre = archivos[0]

        for i in archivos:
            if i ==  dane_des_emp_mensual_nombre:
                try:
                    data = load_workbook(path+"\{}".format(i))
                    sheets = pd.Series(data.sheetnames).str.lower()
                    tnal_mensual_index = sheets[sheets.str.contains('tnal mensual')].index[0]
                except:
                    data = xlrd.open_workbook_xls(path+"\{}".format(i))
                    sheets = pd.Series(data.sheet_names()).str.lower()
                    tnal_mensual_index = sheets[sheets.str.contains('tnal mensual')].index[0]

                df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_mensual_index)

                ind = ['tgp','to','td','ocupados','desocupados','inactivos']
                series = pd.DataFrame({})
                for i in ind:
                    ser_index = df[df.applymap(lambda x: str(x).lower() == i)].dropna(how='all',axis=0).index[0]
                    ser = df.iloc[ser_index,1:].rename(i)
                    if i == 'ocupados' or i == 'desocupados' or i == 'inactivos':
                        ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')*1000      
                    else:
                        ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')/100

                    series[i] = ser

        series.to_csv(path+"\CSV\{}_mensual.csv".format(dane_des_emp_mensual_nombre[:-5]),sep=';',decimal=',')

        shutil.move(path+r'\{}'.format(dane_des_emp_mensual_nombre),path+r'\archivos_fuente\{}'.format(dane_des_emp_mensual_nombre))

    except:
        print('El : {} no se pudo limpiar correctamente'.format(dane_des_emp_mensual_nombre))
        pass


def clean_desempleo_empleo_sexo(path):
    
    try:
        os.mkdir(path+"\\archivos_fuente")
        os.mkdir(path+"\\CSV")
    except:
        pass 
    try:
        archivos =  os.listdir(path)
        archivos.remove("archivos_fuente")
        archivos.remove("CSV")
        dane_sexo_nombre = archivos[0]

        for i in archivos:
            if i == dane_sexo_nombre:
                try: 
                    data = load_workbook(path+"\{}".format(i))
                    sheets = data.sheetnames
                except:
                    data = xlrd.open_workbook_xls(path+"\{}".format(i))
                    sheets = data.sheet_names()

                sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                tnal_nacional_sexo_index = sheets[sheets.str.contains('pytn')].index[0]

                df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_sexo_index)


                ind = [i.lower().replace(' ','_') for i in pd.Series("""% población en edad de trabajar 
TGP
TO
TD
T.D. Abierto
T.D. Oculto
Población total
Población en edad de trabajar
Población económicamente activa
Ocupados
Desocupados
Abiertos
Ocultos
Inactivos""").str.split('\n')[0]]


                series_tnac = pd.DataFrame({})
                series_hombres = pd.DataFrame({})
                series_mujeres = pd.DataFrame({})

                for i in ind:

                    # Total Nacional
                    ser_index_nac = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[0]

                    if ser_index_nac:
                        ser = df.iloc[ser_index_nac,1:].rename(i) 
                        if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')/100
                        else:
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')*1000
                        series_tnac[i] = ser

                    #Hombres
                    ser_index_hom = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[1]

                    if ser_index_hom:
                        ser = df.iloc[ser_index_hom,1:].rename(i)  
                        if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')/100
                        else:
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')*1000
                        series_hombres[i] = ser

                    #Mujeres
                    ser_index_muj = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[2]

                    if ser_index_muj:
                        ser = df.iloc[ser_index_muj,1:].rename(i)  
                        if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')/100
                        else:
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')*1000
                        series_mujeres[i] = ser

        series_tnac.to_csv(path+"\CSV\{}_desempleo_tnac_sexo.csv".format(dane_sexo_nombre[:-5]),sep=';',decimal=',',encoding = 'utf-8')
        series_hombres.to_csv(path+"\CSV\{}_desempleo_hombres.csv".format(dane_sexo_nombre[:-5]),sep=';',decimal=',',encoding = 'utf-8')
        series_mujeres.to_csv(path+"\CSV\{}_desempleo_mujeres.csv".format(dane_sexo_nombre[:-5]),sep=';',decimal=',',encoding = 'utf-8')
        shutil.move(path+r"\{}".format(dane_sexo_nombre),path+r"\archivos_fuente\{}".format(dane_sexo_nombre))
    except:
        print('El : {} no se pudo limpiar correctamente'.format(dane_sexo_nombre))
        pass

    
def clean_desempleo_empleo_regiones(path):        
    try:
        os.mkdir(path+"\\archivos_fuente")
        os.mkdir(path+"\\CSV")
    except:
        pass 
    try:
        archivos =  os.listdir(path)
        archivos.remove("archivos_fuente")
        archivos.remove("CSV")
        dane_regiones_nombre = archivos[0]

        for i in archivos:
            if i == dane_regiones_nombre:
                try: 
                    data = load_workbook(path+"\{}".format(i))
                    sheets = data.sheetnames
                except:
                    data = xlrd.open_workbook_xls(path+"\{}".format(i))
                    sheets = data.sheet_names()

                sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                tnal_nacional_regiones_index = sheets[sheets.str.contains('regionestotalnacional')].index[0]

                df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_regiones_index)


                ind = [i.lower().replace(' ','_') for i in pd.Series("""% población en edad de trabajar 
TGP
TO
TD
T.D. Abierto
T.D. Oculto
Población total
Población en edad de trabajar
Población económicamente activa
Ocupados
Desocupados
Abiertos
Ocultos
Inactivos""").str.split('\n')[0]]


                series_tnac = pd.DataFrame({})
                series_caribe = pd.DataFrame({})
                series_oriental = pd.DataFrame({})
                series_central = pd.DataFrame({})
                series_pacifica = pd.DataFrame({})
                series_bogota = pd.DataFrame({})
                for i in ind:

                    # Total Nacional
                    ser_index_nac = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[0]

                    if ser_index_nac:
                        ser = df.iloc[ser_index_nac,1:].rename(i) 
                        if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')/100
                        else:
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                        series_tnac[i] = ser

                    # Región Caribe
                    ser_index_caribe = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[1]

                    if ser_index_caribe:
                        ser = df.iloc[ser_index_caribe,1:].rename(i)  
                        if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')/100
                        else:
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                        series_caribe[i] = ser

                    #Región oriental
                    ser_index_oriental = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[2]

                    if ser_index_oriental:
                        ser = df.iloc[ser_index_oriental,1:].rename(i)  
                        if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')/100
                        else:
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                        series_oriental[i] = ser

                    #Región Central

                    ser_index_central = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[3]

                    if ser_index_central:
                        ser = df.iloc[ser_index_central,1:].rename(i)  
                        if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')/100
                        else:
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                        series_central[i] = ser

                    #Región pacifica

                    ser_index_pacifica = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[4]

                    if ser_index_pacifica:
                        ser = df.iloc[ser_index_pacifica,1:].rename(i)  
                        if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')/100
                        else:
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                        series_pacifica[i] = ser  

                    #Bogotá

                    ser_index_bogota = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[5]

                    if ser_index_bogota:
                        ser = df.iloc[ser_index_bogota,1:].rename(i)  
                        if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')/100
                        else:
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                        series_bogota[i] = ser
        series_tnac.to_csv(path+"\CSV\{}_desempleo_tnac_regiones.csv".format(dane_regiones_nombre[:-5]),sep=';',decimal=',',encoding = 'utf-8')
        series_caribe.to_csv(path+"\CSV\{}_desempleo_region_caribe.csv".format(dane_regiones_nombre[:-5]),sep=';',decimal=',',encoding = 'utf-8')
        series_oriental.to_csv(path+"\CSV\{}_desempleo_region_oriental.csv".format(dane_regiones_nombre[:-5]),sep=';',decimal=',',encoding = 'utf-8')
        series_central.to_csv(path+"\CSV\{}_desempleo_region_central.csv".format(dane_regiones_nombre[:-5]),sep=';',decimal=',',encoding = 'utf-8')
        series_pacifica.to_csv(path+"\CSV\{}_desempleo_region_pacifica.csv".format(dane_regiones_nombre[:-5]),sep=';',decimal=',',encoding = 'utf-8')
        series_bogota.to_csv(path+"\CSV\{}_desempleo_region_bogota.csv".format(dane_regiones_nombre[:-5]),sep=';',decimal=',',encoding = 'utf-8')
        shutil.move(path+r"\{}".format(dane_regiones_nombre),path+r"\archivos_fuente\{}".format(dane_regiones_nombre))
        
    except:
        print('El : {} no se pudo limpiar correctamente'.format(dane_regiones_nombre))
        pass
    
    
def clean_desempleo_estacionalizado(path):
    try:
        os.mkdir(path+"\\archivos_fuente")
        os.mkdir(path+"\\CSV")
    except:
        pass 
    try:
        archivos =  os.listdir(path)
        archivos.remove("archivos_fuente")
        archivos.remove("CSV")
        dane_des_emp_mensual_nombre = archivos[0]
        
        for i in archivos:
            if i ==  dane_des_emp_mensual_nombre:
                try:
                    data = load_workbook(path+"\{}".format(i))
                    sheets = pd.Series(data.sheetnames).str.lower()
                except:
                    data = xlrd.open_workbook_xls(path+"\{}".format(i))
                    sheets = pd.Series(data.sheet_names()).str.lower()

                # Desempleo estacionalizado mensual
                try:
                    tnal_estacionalizado_mensual_index = sheets[sheets.str.contains('tnal mensual')].index[0]

                    df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_estacionalizado_mensual_index)

                    l = df.iloc[:,0]
                    sup = l[l.str.contains('Concepto').fillna(False)].index[0]


                    df_temp = df.iloc[sup:sup+33,:]
                    df_temp = df_temp.T.dropna(how='all',axis=0).fillna(method='ffill')
                    df_temp.columns = df_temp.iloc[0,:]

                    df_temp.reset_index(level=0,drop=True,inplace=True)
                    df_temp = df_temp.drop([0],axis=0)

                    fecha = pd.date_range(start='2001-01-01',periods=len(df_temp),freq='M',name='Fecha')
                    df_temp.set_index(fecha,drop=True,inplace=True) 
                    df_temp = df_temp.dropna(how='all',axis=1)
                    df_temp = df_temp.iloc[:,2:].applymap(lambda x: float(x))
                    df_temp.columns =  ['% población en edad de trabajar ', 'TGP', 'TO', 'TD', 'T.D. Abierto',
                                            'T.D. Oculto', 'Tasa de subempleo subjetivo',
                                            '  Insuficiencia de horas_1', '  Empleo inadecuado por competencias_1',
                                            '  Empleo inadecuado por ingresos_1', 'Tasa de subempleo objetivo',
                                            '  Insuficiencia de horas_2', '  Empleo inadecuado por competencias_2',
                                            '  Empleo inadecuado por ingresos_2', 'Población total',
                                            'Población en edad de trabajar', 'Población económicamente activa',
                                            'Ocupados', 'Desocupados', 'Abiertos', 'Ocultos', 'Inactivos',
                                            'Subempleados Subjetivos', '  Insuficiencia de horas_3',
                                            '  Empleo inadecuado por competencias_3',
                                            '  Empleo inadecuado por ingresos_3', 'Subempleados Objetivos',
                                            '  Insuficiencia de horas_4', '  Empleo inadecuado por competencias_4',
                                            '  Empleo inadecuado por ingresos_4']
                    
                    
                    tasas_tnac = ['% población en edad de trabajar ', 'TGP', 'TO', 'TD', 'T.D. Abierto',
                                'T.D. Oculto', 'Tasa de subempleo subjetivo',
                                '  Insuficiencia de horas_1', '  Empleo inadecuado por competencias_1',
                                '  Empleo inadecuado por ingresos_1', 'Tasa de subempleo objetivo',
                                '  Insuficiencia de horas_2', '  Empleo inadecuado por competencias_2',
                                '  Empleo inadecuado por ingresos_2',]

                    for colm in df_temp.columns:
                        if colm in tasas_tnac:
                            df_temp.iloc[:][colm] = df_temp.iloc[:][colm]/100
                        else:
                            df_temp.iloc[:][colm] = df_temp.iloc[:][colm]*1000
                    
                    df_temp.to_csv(path+r"\CSV\{}_desempleo_estacionalizado_total_nacional_mensual.csv".format(dane_des_emp_mensual_nombre[:-5]),sep=';',decimal=',',encoding='utf-8')
                except:
                    print("El desempleo estacionalizado total nacional mensual no se pudo limpiar correctamente")
                    pass


                # Desempleo estacionalizado divisiones 
                try:
                    tnal_estacionalizado_divi_index = sheets[sheets.str.contains('tnal cabe ru trim movil')].index[0]

                    df = pd.read_excel(path+r"\{}".format(i),sheet_name=tnal_estacionalizado_divi_index)


                    divisiones = ["Total Nacional","Total Cabeceras","Centros poblados y rural disperso"]

                    for div in divisiones:
                        l = df.iloc[:,0]
                        sup = l[l.str.contains(div).fillna(False)].index[0]


                        df_temp = df.iloc[sup:sup+36,:]
                        df_temp = df_temp.T.dropna(how='all',axis=0).reset_index(level=0,drop=True)
                        df_temp.columns = df_temp.iloc[0,:]

                        df_temp.reset_index(level=0,drop=True,inplace=True)
                        df_temp = df_temp.drop([0],axis=0)

                        fecha = pd.date_range(start='2001-01-01',periods=len(df_temp),freq='M',name='Fecha')

                        df_temp.set_index(fecha,drop=True,inplace=True)
                        df_temp = df_temp.dropna(how='all',axis=1)
                        df_temp = df_temp.iloc[:,2:].applymap(lambda x: float(x))
                        df_temp.columns =  ['% población en edad de trabajar ', 'TGP', 'TO', 'TD', 'T.D. Abierto',
                                            'T.D. Oculto', 'Tasa de subempleo subjetivo',
                                            '  Insuficiencia de horas_1', '  Empleo inadecuado por competencias_1',
                                            '  Empleo inadecuado por ingresos_1', 'Tasa de subempleo objetivo',
                                            '  Insuficiencia de horas_2', '  Empleo inadecuado por competencias_2',
                                            '  Empleo inadecuado por ingresos_2', 'Población total',
                                            'Población en edad de trabajar', 'Población económicamente activa',
                                            'Ocupados', 'Desocupados', 'Abiertos', 'Ocultos', 'Inactivos',
                                            'Subempleados Subjetivos', '  Insuficiencia de horas_3',
                                            '  Empleo inadecuado por competencias_3',
                                            '  Empleo inadecuado por ingresos_3', 'Subempleados Objetivos',
                                            '  Insuficiencia de horas_4', '  Empleo inadecuado por competencias_4',
                                            '  Empleo inadecuado por ingresos_4']
                        
                        
                        tasas_div = ['% población en edad de trabajar ', 'TGP', 'TO', 'TD', 'T.D. Abierto',
                                            'T.D. Oculto', 'Tasa de subempleo subjetivo',
                                            '  Insuficiencia de horas_1', '  Empleo inadecuado por competencias_1',
                                            '  Empleo inadecuado por ingresos_1', 'Tasa de subempleo objetivo',
                                            '  Insuficiencia de horas_2', '  Empleo inadecuado por competencias_2',
                                            '  Empleo inadecuado por ingresos_2',]
                        
                        for colm in df_temp.columns:
                            if colm in tasas_div:
                                df_temp.iloc[:][colm] = df_temp.iloc[:][colm]/100
                            else:
                                df_temp.iloc[:][colm] = df_temp.iloc[:][colm]*1000
                        df_temp.to_csv(path+r"\CSV\{}_desempleo_estacionalizado_divisiones_{}.csv".format(dane_des_emp_mensual_nombre[:-5],div),sep=';',decimal=',',encoding='utf-8')
                except:
                    print("El desempleo estacionalizado por divisiones no se pudo limpier correctamente")
                    pass

                # Desempleo estacionalizado por areas
                try:
                    tnal_estacionalizado_areas_index = sheets[sheets.str.contains('areas trim movil')].index[0]

                    df = pd.read_excel(path+"\{}".format(i),tnal_estacionalizado_areas_index)

                    l = df.iloc[:,0]
                    sup = l[l.str.contains('Total 13 ciudades y áreas metropolitanas').fillna(False)].index[-1]
                    inf = l[l.str.contains('Total 23 ciudades y A.M.').fillna(False)].index[-1]

                    df = df.iloc[sup:inf+28,:]
                    df.reset_index(level=0,drop=True,inplace=True)

                    ind = [i.lower().replace(' ','_') for i in pd.Series(['% población en edad de trabajar ',
                    'TGP',
                    'TO',
                    'TD',
                    'T.D. Abierto',
                    'T.D. Oculto',
                    'Población total',
                    'Población en edad de trabajar',
                    'Población económicamente activa',
                    'Ocupados',
                    'Desocupados',
                    'Abiertos',
                    'Ocultos',
                    'Inactivos',]).str.split(',').str[0]]

                    ciudades = pd.Series(['Total 13 ciudades y áreas metropolitanas',
                        'Bogotá',
                        'Medellín A.M.',
                        'Cali A.M.',
                        'Barranquilla A.M.',
                        'Bucaramanga A.M.',
                        'Manizales A.M.',
                        'Pasto',
                        'Pereira A.M.',
                        'Cúcuta A.M.',
                        'Ibagué',
                        'Montería',
                        'Cartagena',
                        'Villavicencio',
                        'Tunja',
                        'Florencia',
                        'Popayán',
                        'Valledupar',
                        'Quibdó',
                        'Neiva',
                        'Riohacha',
                        'Santa Marta',
                        'Armenia',
                        'Sincelejo',
                        'Total 10 ciudades',
                        'Total 23 ciudades y A.M.']).str.upper()

                    ser_index_name = df[df.applymap(lambda x: str(x).lower().replace(' ','_')=='concepto')].dropna(how='all',axis=0).index[:]

                    dic_ini = pd.DataFrame({})
                    contador = 0
                    for j in ind:
                        ser_index_nac = df[df.applymap(lambda x: str(x).lower().replace(' ','_')==j)].dropna(how='all',axis=0).index[:]
                        for jx in ser_index_name:
                            ser_ = df.iloc[ser_index_nac,0:]
                            ser_ = ser_.reset_index(level=0,drop=True)
                            ser_.insert(0,'Ciudad',ciudades)

                            contador += 1
                            if contador == 1:
                                dic_ini = ser_

                        dic_ini = pd.concat([dic_ini,ser_],axis=0)

                    fecha = pd.date_range(start="2001-01-01",freq='M',periods=len(dic_ini.columns[2:]),name='Fecha')
                    dic_ini.columns = ['Ciudad','Indicador'] + [i.date() for i in fecha]
                    dic_ini = dic_ini.drop_duplicates()
                    dic_ini = dic_ini.set_index(['Ciudad','Indicador'],drop=True)
                    dic_ini = dic_ini.applymap(lambda x: str(x).replace('-','0'))
                    dic_ini = dic_ini.applymap(lambda x: float(x))
                    
                    tasas_areas = ['% población en edad de trabajar ','TGP','TO','TD','T.D. Abierto','T.D. Oculto']
                    
                    for mulindex in dic_ini.index:
                        if mulindex[1] in tasas_areas:
                            dic_ini.loc[(mulindex[0],mulindex[1])] = dic_ini.loc[(mulindex[0],mulindex[1])]/100
                        else:
                            dic_ini.loc[(mulindex[0],mulindex[1])] = dic_ini.loc[(mulindex[0],mulindex[1])]*1000
                                                                            
                    dic_fin = pd.DataFrame()
                    contador_2 = 0

                    for ciu in ciudades:
                        group = dic_ini.groupby(level='Ciudad').get_group(ciu).T
                        group.columns = pd.Series([i[1] for i in group.columns]).str.replace(',','')
                        group = group.applymap(lambda x: float(x))
                        group.insert(0,'Ciudad',ciu)

                        contador_2 +=1

                        if contador_2 == 1:
                            dic_fin = group
                        else:   
                            dic_fin = pd.concat([dic_fin,group]) 
                            
                    dic_fin = dic_fin.rename_axis('Fecha')
                    dic_fin.to_csv(path+r"\CSV\{}_desempleo_estacionalizado_areas_ciudades.csv".format(dane_des_emp_mensual_nombre[:-5]),sep=';',decimal=',',encoding='utf-8')
                except:
                    print("El desempleo estacionalizado por areas no se pudo limpier correctamente")
                    pass

                #Desempleo estacionalizado por ramas ciiu4
                try:
                    tnal_estacionalizado_ramas_index = sheets[sheets.str.contains('ocup ramas trim tnal ciiu 4 ')].index[0]

                    df = pd.read_excel(path+r"\{}".format(i),sheet_name=tnal_estacionalizado_ramas_index)


                    divisiones = ["TOTAL NACIONAL","CABECERAS","CENTROS POBLADOS Y RURAL DISPERSO"]

                    for j in divisiones:
                        l = df.iloc[:,0]
                        sup = l[l.str.contains(j).fillna(False)].index[0]


                        df_temp = df.iloc[sup:sup+20,:]
                        df_temp = df_temp.T.dropna(how='all',axis=0).reset_index(level=0,drop=True)
                        df_temp.columns = df_temp.iloc[0,:]

                        df_temp.reset_index(level=0,drop=True,inplace=True)
                        df_temp = df_temp.drop([0],axis=0)

                        fecha = pd.date_range(start='2015-01-01',periods=len(df_temp),freq='M',name='Fecha')

                        df_temp.set_index(fecha,drop=True,inplace=True)
                        df_temp = df_temp.iloc[:,4:].applymap(lambda x: float(x)*1000)

                        df_temp.to_csv(path+r"\CSV\{}_desempleo_estacionalizado_ramasciiu4_{}.csv".format(dane_des_emp_mensual_nombre[:-5],j),sep=';',decimal=',',encoding='utf-8')
                except:
                    print("El desempleo estacionalizado por ramas ciiu4 no se pudo limpier correctamente")
                    pass

                ## Desempleo estacionalizado por areas y ciiu (ramas y ciudades)
                try:
                    tnal_estacionalizado_areasciiu_index = sheets[sheets.str.contains('ocu ramas trim 23 áreas ciiu 4')].index[0]

                    df = pd.read_excel(path+"\{}".format(i),tnal_estacionalizado_areasciiu_index)

                    l = df.iloc[:,0]
                    sup = l[l.str.contains('OCUPADOS 13 CIUDADES Y ÁREAS METROPOLITANAS').fillna(False)].index[-1]
                    inf = l[l.str.contains('SINCELEJO').fillna(False)].index[-1]

                    df = df.iloc[sup:inf+20,:]
                    df.reset_index(level=0,drop=True,inplace=True)

                    ind = pd.Series(['no_informa',
                    'agricultura,_ganadería,_caza,_silvicultura_y_pesca',
                    'explotación_de_minas_y_canteras',
                    'industrias_manufactureras',
                    'suministro_de_electricidad_gas,_agua_y_gestión_de_desechos',
                    'construcción',
                    'comercio_y_reparación_de_vehículos',
                    'alojamiento_y_servicios_de_comida',
                    'transporte_y_almacenamiento',
                    'información_y_comunicaciones',
                    'actividades_financieras_y_de_seguros',
                    'actividades_inmobiliarias',
                    'actividades_profesionales,_científicas,_técnicas_y_servicios_administrativos',
                    'administración_pública_y_defensa,_educación_y_atención_de_la_salud_humana',
                    'actividades_artísticas,_entretenimiento,_recreación_y_otras_actividades_de_servicios',]).str.replace(',','')

                    ciudades = pd.Series(['OCUPADOS 13 CIUDADES Y ÁREAS METROPOLITANAS',
                    'MEDELLÍN A.M.',
                    'BARRANQUILLA A.M.',
                    'BOGOTÁ',
                    'CARTAGENA',
                    'MANIZALES A.M.',
                    'MONTERÍA',
                    'VILLAVICENCIO',
                    'PASTO',
                    'CÚCUTA A.M.',
                    'PEREIRA A.M.',
                    'BUCARAMANGA A.M.',
                    'IBAGUÉ',
                    'CALI  A.M.',
                    'TUNJA',
                    'FLORENCIA',
                    'POPAYÁN',
                    'VALLEDUPAR',
                    'QUIBDÓ',
                    'NEIVA',
                    'RIOHACHA',
                    'SANTA MARTA',
                    'ARMENIA',
                    'SINCELEJO']).str.upper()

                    ser_index_name = df[df.applymap(lambda x: str(x).lower().replace(' ','_').replace(',','')=='concepto')].dropna(how='all',axis=0).index[:]

                    dic_ini = pd.DataFrame({})
                    contador = 0
                    for j in ind:
                        ser_index_nac = df[df.applymap(lambda x: str(x).lower().replace(' ','_').replace(',','')==j)].dropna(how='all',axis=0).index[:]
                        for jx in ser_index_name:
                            ser_ = df.iloc[ser_index_nac,0:]
                            ser_ = ser_.reset_index(level=0,drop=True)
                            ser_.insert(0,'Ciudad',ciudades)

                            contador += 1
                            if contador == 1:
                                dic_ini = ser_

                        dic_ini = pd.concat([dic_ini,ser_],axis=0)

                    fecha = pd.date_range(start="2015-01-01",freq='M',periods=len(dic_ini.columns[2:]),name='Fecha')
                    dic_ini.columns = ['Ciudad','Indicador'] + [i.date() for i in fecha]
                    dic_ini = dic_ini.drop_duplicates()
                    dic_ini = dic_ini.set_index(['Ciudad','Indicador'],drop=True)
                    dic_ini = dic_ini.applymap(lambda x: str(x).replace('-','0'))
                    dic_ini = dic_ini.applymap(lambda x: float(x)*1000)
                    
                    dic_fin = pd.DataFrame()
                    contador_2 = 0

                    for ciu in ciudades:
                        group = dic_ini.groupby(level='Ciudad').get_group(ciu).T
                        group.columns = pd.Series([i[1] for i in group.columns]).str.replace(',','')
                        group = group.applymap(lambda x: float(x))
                        group.insert(0,'Ciudad',ciu)

                        contador_2 +=1

                        if contador_2 == 1:
                            dic_fin = group
                        else:   
                            dic_fin = pd.concat([dic_fin,group])
                    
                    dic_fin = dic_fin.rename_axis('Fecha')
                    dic_fin.to_csv(path+r"\CSV\{}_ocupados_estacionalizado_areasciiu4_ciudadesyramas.csv".format(dane_des_emp_mensual_nombre[:-5]),sep=';',decimal=',',encoding='utf-8')
                except:
                    print("El desempleo estacionalizado por ramas ciiu4 y ciudades no se pudo limpier correctamente")
                    pass
                
        shutil.move(path+r"\{}".format(dane_des_emp_mensual_nombre),path+r"\archivos_fuente\{}".format(dane_des_emp_mensual_nombre))
    except:
        print('El : {} no se pudo limpiar correctamente'.format(dane_des_emp_mensual_nombre))
        pass