import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

import os
import xlrd
import shutil

class limpieza_mercado_laboral():
    
    def clean_informalidad(self,path):
        try:
            os.mkdir(path+"\\archivos_fuente")
        except:
            pass 
        try:
            archivos =  pd.Series(os.listdir(path))
            dane_informalidad_nombre = archivos[archivos.str.contains('informalidad')].values[0]
            for i in archivos:
                if i ==  dane_informalidad_nombre:
                    try:
                        data = load_workbook(path+"\{}".format(i))
                        sheets = data.sheetnames
                    except:
                        data = xlrd.open_workbook_xls(path+"\{}".format(i))
                        sheets = sheets = data.sheet_names()


                    #Prop informalidad Total Nacional
                    try:
                        sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                        tnal_nacional_informalidad_index = sheets[sheets.str.contains('informalidad')].index[0]

                        df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_informalidad_index)

                        index = df[df.iloc[:,0].str.contains('23 Ciudades').fillna(False)].index[0]

                        data_23_ciudades = df.iloc[index:,:]

                        periodo = data_23_ciudades[data_23_ciudades.iloc[:,1].str.contains('Ene-').fillna(False)].dropna(axis=1).T
                        periodo.columns = ['Periodo']
                        periodo.reset_index(level=0,drop=True,inplace = True)

                        fecha = pd.date_range(start='2007',freq='M',periods=len(periodo),name = 'Fecha')
                        periodo['Fecha'] = fecha

                        tasa = data_23_ciudades[data_23_ciudades.iloc[:,0].str.contains('Ocupados').fillna(False)].dropna(axis=1).T.iloc[1:,:]

                        tasa.columns = ['Ocupados Informales']
                        tasa['Ocupados Informales'] = tasa['Ocupados Informales'].astype('float')
                        tasa.reset_index(level=0,drop=True,inplace=True)

                        tasa_informalidad = pd.concat([periodo,tasa],axis=1)
                        tasa_informalidad.to_csv(path+'\informalidad_total_Nacional.csv',sep=';',decimal=',',index=False)
                    except:
                        print('La propoción de informalidad total nacional no se pudo limpiar correctamente')
                        pass


                    #Informalidad Ciudad
                    try:
                        sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                        tnal_nacional_ciudad_index = sheets[sheets.str.contains('ciudades')].index[0]

                        df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_ciudad_index)

                        l = df.iloc[:,0]
                        sup = l[l.str.contains('Total 13 ciudades y AM').fillna(False)].index[-1]
                        inf = l[l.str.contains('23 ciudades y áreas').fillna(False)].index[-1]

                        df = df.iloc[sup:inf+5,:]

                        df.reset_index(level=0,drop=True,inplace=True)

                        ind = [i.lower().replace(' ','_') for i in pd.Series("""Ocupados
                        Formales
                        Informales""").str.split('\n')[0]]

                        dic = pd.DataFrame({})
                        for j in ind:
                            # Total Nacional
                            ser_index_nac = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == 'informales')].dropna(how='all',axis=0).index
                        for ix in ser_index_nac:
                            ser_ = df.iloc[ix-4:ix+1,0:]
                            ser_ = ser_.T
                            ser_.columns = ser_.iloc[0,:]
                            ser_ = ser_.reset_index(level=0,drop=True).drop([0],axis=0)
                            fecha = pd.date_range(start='2007-01-01',periods=len(ser_),freq='M',name='Fecha')
                            ser_ = ser_.set_index(fecha,drop=True)
                            ser_['Ocupados'] = ser_['Ocupados'].astype('float')*1000
                            ser_['Formales'] = ser_['Formales'].astype('float')*1000
                            ser_['Informales'] = ser_['Informales'].astype('float')*1000

                            ser_ = ser_.rename(columns={'Ocupados':'Ocupados_{}'.format(ser_.columns[0][:3]),
                                                        'Formales':'Formales_{}'.format(ser_.columns[0][:3]),
                                                        'Informales':'Informales_{}'.format(ser_.columns[0][:3])
                                                        })

                            ser_ = ser_.fillna(method='ffill')

                            dic['Trimestre Móvil'] = ser_.iloc[:,1]

                            dic[ser_.columns[0]] = ser_.iloc[:,0].apply(lambda x: str(x).replace(str(x),ser_.columns[0]))
                            dic[ser_.columns[2]] = ser_.iloc[:,2].apply(lambda x: float(x))
                            dic[ser_.columns[3]] = ser_.iloc[:,3].apply(lambda x: float(x))
                            dic[ser_.columns[4]] = ser_.iloc[:,4].apply(lambda x: float(x))

                        dic.to_csv(path+r"\informalidad_ciudades.csv",sep=';',decimal=',',encoding='utf-8')
                    except:
                        print("La informalidad por Ciudades no se pudo limpiar correctamente")
                        pass

                    #Informalidad por sexo
                    try:
                        sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                        tnal_nacional_sexo_index = sheets[sheets.str.contains('sexo')].index[0]
                        df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_sexo_index)

                        ciudades = ["Ocupados 13 ciudades y áreas metropolitanas","Ocupados 23 ciudades y áreas metropolitanas"]

                        for j in ciudades:
                            l = df.iloc[:,0]
                            sup = l[l.str.contains(j).fillna(False)].index[0]

                            df_temp = df.iloc[sup:sup+9,:]
                            df_temp = df_temp.T.dropna(how='all',axis=0)
                            df_temp.columns = df_temp.iloc[0,:]

                            df_temp.reset_index(level=0,drop=True,inplace=True)
                            df_temp = df_temp.drop([0],axis=0)

                            fecha = pd.date_range(start='2007-01-01',periods=len(df_temp),freq='M',name='Fecha')

                            df_temp.set_index(fecha,drop=True,inplace=True)
                            df_temp = df_temp.applymap(lambda x: float(x)*1000)

                            df_temp.to_csv(path+r"\informalidad_sexo_{}.csv".format(j),sep=';',decimal=',',encoding='utf-8')
                    except:
                        print("La informalidad por sexo no se pudo limpiar correctamente")
                        pass
                    ## Imformalidad por educacion

                    try:
                        sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                        tnal_nacional_educacion_index = sheets[sheets.str.contains('educación')].index[0]
                        df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_educacion_index)

                        ciudades = ["Ocupados 13 ciudades y áreas metropolitanas","Ocupados 23 ciudades y áreas metropolitanas"]

                        for j in ciudades:
                            l = df.iloc[:,0]
                            sup = l[l.str.contains(j).fillna(False)].index[0] 
                            df_temp = df.iloc[sup:sup+18,:]
                            df_temp = df_temp.T.dropna(how='all',axis=0)
                            df_temp.columns = df_temp.iloc[0,:]

                            df_temp.reset_index(level=0,drop=True,inplace=True)
                            df_temp = df_temp.drop([0],axis=0)

                            fecha = pd.date_range(start='2007-01-01',periods=len(df_temp),freq='M',name='Fecha')

                            df_temp.set_index(fecha,drop=True,inplace=True)
                            df_temp = df_temp.applymap(lambda x: float(x)*1000)

                            df_temp.to_csv(path+r"\informalidad_educacion_{}.csv".format(j),sep=';',decimal=',',encoding='utf-8')

                    except:
                        print("La informalidad por educacion no se pudo limpiar correctamente")
                        pass


                    ## Informalidad por ramas ciiu 4a
                    try:
                        sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                        tnal_nacional_educacion_index = sheets[sheets.str.contains('ciiu4')].index[0]
                        df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_educacion_index)

                        ciudades = ["Total 13 áreas","Ocupados 23 ciudades y áreas metropolitanas"]

                        for j in ciudades:
                            l = df.iloc[:,0]
                            sup = l[l.str.contains(j).fillna(False)].index[0] 
                            df_temp = df.iloc[sup:sup+48,:]
                            df_temp = df_temp.T.dropna(how='all',axis=0)
                            df_temp.columns = df_temp.iloc[0,:]

                            df_temp.reset_index(level=0,drop=True,inplace=True)
                            df_temp = df_temp.drop([0],axis=0)

                            fecha = pd.date_range(start='2007-01-01',periods=len(df_temp),freq='M',name='Fecha')

                            df_temp.set_index(fecha,drop=True,inplace=True)
                            df_temp = df_temp.applymap(lambda x: float(x)*1000)

                            df_temp.to_csv(path+r"\informalidad_ramasciiu4a_{}.csv".format(j),sep=';',decimal=',',encoding='utf-8')
                    except:
                        print("La informalidad por ramas CIIU4a no se pudo limpiar corectamente")
                        pass

                    #Informalidad por Seguridad Social
                    try:
                        sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                        tnal_nacional_educacion_index = sheets[sheets.str.contains('seguridadsocial13')].index[0]
                        df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_educacion_index)

                        ciudades = ["Ocupados 13 ciudades y áreas metropolitanas","Ocupados 23 ciudades y áreas metropolitanas"]


                        #Cantidad de personas
                        for j in ciudades:
                            l = df.iloc[:,0]
                            sup = l[l.str.contains(j).fillna(False)].index[0] 
                            df_temp = df.iloc[sup:sup+10,:]
                            df_temp = df_temp.T.dropna(how='all',axis=0)
                            df_temp.columns = df_temp.iloc[0,:]

                            df_temp.reset_index(level=0,drop=True,inplace=True)
                            df_temp = df_temp.drop([0],axis=0)

                            fecha = pd.date_range(start='2007-01-01',periods=len(df_temp),freq='M',name='Fecha')

                            df_temp.set_index(fecha,drop=True,inplace=True)
                            df_temp = df_temp.applymap(lambda x: float(x)*1000)

                            df_temp.to_csv(path+r"\informalidad_segsocial_cantidad_{}.csv".format(j),sep=';',decimal=',',encoding='utf-8')

                        #Porcentaje de personas
                        for j in ciudades:
                            l = df.iloc[:,0]
                            sup = l[l.str.contains(j).fillna(False)].index[1] 
                            df_temp = df.iloc[sup:sup+10,:]
                            df_temp = df_temp.T.dropna(how='all',axis=0)
                            df_temp.columns = df_temp.iloc[0,:]

                            df_temp.reset_index(level=0,drop=True,inplace=True)
                            df_temp = df_temp.drop([0],axis=0)

                            fecha = pd.date_range(start='2007-01-01',periods=len(df_temp),freq='M',name='Fecha')

                            df_temp.set_index(fecha,drop=True,inplace=True)
                            df_temp = df_temp.applymap(lambda x: float(x))

                            df_temp.to_csv(path+r"\informalidad_segsocial_porcentaje_{}.csv".format(j),sep=';',decimal=',',encoding='utf-8')
                    except:
                        print("La informalidad por seguridad social no se pudo limpiar correctamente")
                        pass

            shutil.move(path+r'\{}'.format(dane_informalidad_nombre),path+r'\archivos_fuente\{}'.format(dane_informalidad_nombre))
        except:
            pass

    def clean_desempleo_empleo_mensual(self,path):
        
        try:
            os.mkdir(path+"\\archivos_fuente")
        except:
            pass 
        
        try:
            archivos = pd.Series(os.listdir(path))
            dane_des_emp_mensual_nombre = archivos[archivos.str.contains('anexo_desestacionalizado_empleo')].values[0]

            for i in archivos:
                if i ==  dane_des_emp_mensual_nombre:
                    try:
                        data = load_workbook(path+"\{}".format(i))
                        sheets = pd.Series(data.sheetnames).str.lower()
                        tnal_mensual_index = sheets[sheets.str.contains('tnal mensual')].index[0]
                    except:
                        data = xlrd.open_workbook_xls(path+"\{}".format(i))
                        sheets = pd.Series(data.sheet_names()).str.lower()
                        tnal_mensual_index = sheets[sheets.str.contains('tnal mensual')].index[0]

                    df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_mensual_index)

                    ind = ['tgp','to','td','ocupados','desocupados','inactivos']
                    series = pd.DataFrame({})
                    for i in ind:
                        ser_index = df[df.applymap(lambda x: str(x).lower() == i)].dropna(how='all',axis=0).index[0]
                        ser = df.iloc[ser_index,1:].rename(i)
                        if i == 'ocupados' or i == 'desocupados' or i == 'inactivos':
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')*1000      
                        else:
                            ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')

                        series[i] = ser

            series.to_csv(path+"\desempleo_desest_mensual.csv",sep=';',decimal=',')

            shutil.move(path+r'\{}'.format(dane_des_emp_mensual_nombre),path+r'\archivos_fuente\{}'.format(dane_des_emp_mensual_nombre))

        except:
            print('El : {} no se pudo limpiar correctamente'.format(dane_des_emp_mensual_nombre))
            pass


    def clean_desempleo_empleo_sexo(self,path):
        
        try:
            os.mkdir(path+"\\archivos_fuente")
        except:
            pass      

        try:
            archivos = pd.Series(os.listdir(path))
            dane_sexo_nombre = archivos[archivos.str.contains('anexo_sexo_')].values[0]

            for i in archivos:
                if i == dane_sexo_nombre:
                    try: 
                        data = load_workbook(path+"\{}".format(i))
                        sheets = data.sheetnames
                    except:
                        data = xlrd.open_workbook_xls(path+"\{}".format(i))
                        sheets = data.sheet_names()

                    sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                    tnal_nacional_sexo_index = sheets[sheets.str.contains('pytn')].index[0]

                    df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_sexo_index)


                    ind = [i.lower().replace(' ','_') for i in pd.Series("""% población en edad de trabajar 
TGP
TO
TD
T.D. Abierto
T.D. Oculto
Población total
Población en edad de trabajar
Población económicamente activa
Ocupados
Desocupados
Abiertos
Ocultos
Inactivos""").str.split('\n')[0]]


                    series_tnac = pd.DataFrame({})
                    series_hombres = pd.DataFrame({})
                    series_mujeres = pd.DataFrame({})

                    for i in ind:

                        # Total Nacional
                        ser_index_nac = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[0]

                        if ser_index_nac:
                            ser = df.iloc[ser_index_nac,1:].rename(i) 
                            if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')
                            else:
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')*1000
                            series_tnac[i] = ser

                        #Hombres
                        ser_index_hom = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[1]

                        if ser_index_hom:
                            ser = df.iloc[ser_index_hom,1:].rename(i)  
                            if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')
                            else:
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')*1000
                            series_hombres[i] = ser

                        #Mujeres
                        ser_index_muj = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[2]

                        if ser_index_muj:
                            ser = df.iloc[ser_index_muj,1:].rename(i)  
                            if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')
                            else:
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='M',name = 'Fecha')).astype('float')*1000
                            series_mujeres[i] = ser

            series_tnac.to_csv(path+"\desempleo_tnac_sexo.csv",sep=';',decimal=',',encoding = 'utf-8')
            series_hombres.to_csv(path+"\desempleo_hombres.csv",sep=';',decimal=',',encoding = 'utf-8')
            series_mujeres.to_csv(path+"\desempleo_mujeres.csv",sep=';',decimal=',',encoding = 'utf-8')
            shutil.move(path+r"\{}".format(dane_sexo_nombre),path+r"\archivos_fuente\{}".format(dane_sexo_nombre))
        except:
            print('El : {} no se pudo limpiar correctamente'.format(dane_sexo_nombre))
            pass

        
    def clean_desempleo_empleo_regiones(self,path):        
        try:
            os.mkdir(path+"\\archivos_fuente")
        except:
            pass 
        try:
            archivos = pd.Series(os.listdir(path))
            dane_regiones_nombre = archivos[archivos.str.contains('anexo_ech_regiones')].values[0]

            for i in archivos:
                if i == dane_regiones_nombre:
                    try: 
                        data = load_workbook(path+"\{}".format(i))
                        sheets = data.sheetnames
                    except:
                        data = xlrd.open_workbook_xls(path+"\{}".format(i))
                        sheets = data.sheet_names()

                    sheets = pd.Series(sheets).str.lower().str.replace(' ','')
                    tnal_nacional_regiones_index = sheets[sheets.str.contains('regionestotalnacional')].index[0]

                    df = pd.read_excel(path+"\{}".format(i),sheet_name=tnal_nacional_regiones_index)


                    ind = [i.lower().replace(' ','_') for i in pd.Series("""% población en edad de trabajar 
TGP
TO
TD
T.D. Abierto
T.D. Oculto
Población total
Población en edad de trabajar
Población económicamente activa
Ocupados
Desocupados
Abiertos
Ocultos
Inactivos""").str.split('\n')[0]]


                    series_tnac = pd.DataFrame({})
                    series_caribe = pd.DataFrame({})
                    series_oriental = pd.DataFrame({})
                    series_central = pd.DataFrame({})
                    series_pacifica = pd.DataFrame({})
                    series_bogota = pd.DataFrame({})
                    for i in ind:

                        # Total Nacional
                        ser_index_nac = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[0]

                        if ser_index_nac:
                            ser = df.iloc[ser_index_nac,1:].rename(i) 
                            if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')
                            else:
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                            series_tnac[i] = ser

                        # Región Caribe
                        ser_index_caribe = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[1]

                        if ser_index_caribe:
                            ser = df.iloc[ser_index_caribe,1:].rename(i)  
                            if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')
                            else:
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                            series_caribe[i] = ser

                        #Región oriental
                        ser_index_oriental = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[2]

                        if ser_index_oriental:
                            ser = df.iloc[ser_index_oriental,1:].rename(i)  
                            if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')
                            else:
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                            series_oriental[i] = ser

                        #Región Central

                        ser_index_central = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[3]

                        if ser_index_central:
                            ser = df.iloc[ser_index_central,1:].rename(i)  
                            if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')
                            else:
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                            series_central[i] = ser

                        #Región pacifica

                        ser_index_pacifica = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[4]

                        if ser_index_pacifica:
                            ser = df.iloc[ser_index_pacifica,1:].rename(i)  
                            if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')
                            else:
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                            series_pacifica[i] = ser  

                        #Bogotá

                        ser_index_bogota = df[df.applymap(lambda x: str(x).lower().replace(' ','_') == i)].dropna(how='all',axis=0).index[5]

                        if ser_index_bogota:
                            ser = df.iloc[ser_index_bogota,1:].rename(i)  
                            if i=='%_población_en_edad_de_trabajar_' or i == 'tgp' or i== 'to' or i == 'td' or i == 't.d._abierto' or i == 't.d._oculto':
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')
                            else:
                                ser = ser.set_axis(pd.date_range(start='2001-01-01',periods=len(ser),freq='6M',name = 'Fecha')).astype('float')*1000
                            series_bogota[i] = ser
            series_tnac.to_csv(path+"\desempleo_tnac_regiones.csv",sep=';',decimal=',',encoding = 'utf-8')
            series_caribe.to_csv(path+"\desempleo_region_caribe.csv",sep=';',decimal=',',encoding = 'utf-8')
            series_oriental.to_csv(path+"\desempleo_region_oriental.csv",sep=';',decimal=',',encoding = 'utf-8')
            series_central.to_csv(path+"\desempleo_region_central.csv",sep=';',decimal=',',encoding = 'utf-8')
            series_pacifica.to_csv(path+"\desempleo_region_pacifica.csv",sep=';',decimal=',',encoding = 'utf-8')
            series_bogota.to_csv(path+"\desempleo_region_bogota.csv",sep=';',decimal=',',encoding = 'utf-8')
            shutil.move(path+r"\{}".format(dane_regiones_nombre),path+r"\archivos_fuente\{}".format(dane_regiones_nombre))
            
        except:
            print('El : {} no se pudo limpiar correctamente'.format(dane_regiones_nombre))
            pass