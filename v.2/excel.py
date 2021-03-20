import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

import os
import shutil


def guardar_excel(Fuente,carpeta_origen,nombre_archivo,carpeta_destino,hyperlinks=False):
    wb =  Workbook()
    ws = wb.active
    ws.title = 'Indice'
    ws.sheet_properties.tabColor = '51B6DC'

    fuente = wb.create_sheet('Fuente',1)
    for r in dataframe_to_rows(Fuente,index=False):
        fuente.append(r)

    
    archivos = os.listdir(carpeta_origen)
    
    for ix,ind in enumerate(archivos):
        
        try:
            data = pd.read_csv(carpeta_origen+"\{}".format(ind),sep=';',decimal=',')
        except:
            continue
        
        # Encabezados y valores
        header = tuple(data.columns)
        values = [tuple(v) for v in data.values]
        
        if len(ind) > 30:
            ind2 = ind[:27]+str(np.random.randint(100,999,1)[0])
            sheet_name = ind2.replace(' ','_').replace(',','')
        else:
            sheet_name = ind[:-4].replace(' ','_').replace(',','')
        
        wss = wb.create_sheet("{}".format(sheet_name))
        wss.append(header)
        
        for val in values:
            wss.append(val)
            
        if hyperlinks:
                
            
            link = "{}.xlsx#{}!F1".format(nombre_archivo,sheet_name)
            name = '{}'.format(sheet_name)

            ws.cell(row=ix+2, column=2).value = '=HYPERLINK("{}", "{}")'.format(link,name)
            ws.cell(row=ix+2, column=2).style = 'Hyperlink'
            ws.cell(row=ix+2, column=4).value = ind[:-4]
            
            ix_col = len(data.columns)+2
            
            wss.cell(row=1,column=ix_col).value = '=HYPERLINK("{}", "{}")'.format("{}.xlsx#Indice!A1".format(nombre_archivo),"Índice")
            wss.cell(row=1,column=ix_col).style = 'Hyperlink'


    wb.save(carpeta_destino+"\{}.xlsx".format(nombre_archivo))
    os.startfile(carpeta_destino+"\{}.xlsx".format(nombre_archivo),'open')
    
    return ("El archivo {}, se cuardo exitosamente en {}".format(nombre_archivo,carpeta_destino))