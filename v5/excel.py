import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

import os
import shutil


def guardar_excel(Fuente,carpeta_origen,nombre_archivo,carpeta_destino,hyperlinks=False):
    wb =  Workbook()
    ws = wb.active
    ws.title = 'Indice'
    ws.sheet_properties.tabColor = '51B6DC'

    fuente = wb.create_sheet('Fuente',1)
    for r in dataframe_to_rows(Fuente,index=False):
        fuente.append(r)

    
    archivos = os.listdir(carpeta_origen)
    
    for ix,ind in enumerate(archivos):
        
        try:
            data = pd.read_csv(carpeta_origen+"\{}".format(ind),sep=';',decimal=',')
        except:
            continue
        
        # Encabezados y valores
        header = tuple(data.columns)
        values = [tuple(v) for v in data.values]
        
        sheet_name = ind[:7]
        
        wss = wb.create_sheet("{}".format(sheet_name))
        wss.append(header)
        
        for val in values:
            wss.append(val)
            
        if hyperlinks:
                
            
            link = "{}.xlsx#{}!F1".format(nombre_archivo,sheet_name)
            name = '{}'.format(sheet_name)

            ws.cell(row=ix+2, column=2).value = '=HYPERLINK("{}", "{}")'.format(link,name)
            ws.cell(row=ix+2, column=2).style = 'Hyperlink'
            ws.cell(row=ix+2, column=4).value = ind[:-4]
            
            ix_col = len(data.columns)+2
            
            wss.cell(row=1,column=ix_col).value = '=HYPERLINK("{}", "{}")'.format("{}.xlsx#Indice!A1".format(nombre_archivo),"Índice")
            wss.cell(row=1,column=ix_col).style = 'Hyperlink'
            
        
        shutil.move(carpeta_origen+r"\{}".format(ind),carpeta_origen+r"\{}.csv".format(ind[:7]))


    wb.save(carpeta_destino+"\{}.xlsx".format(nombre_archivo))
    os.startfile(carpeta_destino+"\{}.xlsx".format(nombre_archivo),'open')
    
    return ("El archivo {}, se cuardo exitosamente en {}".format(nombre_archivo,carpeta_destino))